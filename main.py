import glob
import openpyxl
from googletrans import Translator

extension = 'xlsx'
excel_files = [i for i in glob.glob('*.{}'.format(extension))]

t = Translator()

for file_name in excel_files:
    workbook = openpyxl.load_workbook(file_name)
    for sheet in workbook.worksheets:
        print(f"======= {sheet} =======")

        if sheet.max_column:
            for idx, val in enumerate(sheet):
                try:
                    if idx == 0 or not val[1].value or sheet.cell(idx + 1, 4).value:
                        continue
                    if val[0].value is None:
                        break

                    kr = val[1].value
                    ja = t.translate(val[1].value, dest='ja').text
                    print(f"{kr} -> {ja}")
                    sheet.cell(idx + 1, 4).value = ja

                except ValueError:
                    workbook.save(file_name)
                    print("an error occurred")
                    break

        workbook.save(file_name)
    print("translate completed!")


